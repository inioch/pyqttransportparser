import sys
import os
import datetime
import csv
import requests
import openpyxl
from openpyxl.styles import Font

# PyQt6 Imports
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QTableWidget,
                             QTableWidgetItem, QHeaderView, QFileDialog,
                             QMessageBox, QMenu, QInputDialog, QAbstractItemView,
                             QFrame)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor

# --- TWOJA KONFIGURACJA FIREBASE ---
FIREBASE_URL = "https://logistykaapp-11e03-default-rtdb.europe-west1.firebasedatabase.app/"


# --- KLASA DO POPRAWNEGO SORTOWANIA LICZB (KLUCZ DO SUKCESU) ---
class NumericItem(QTableWidgetItem):
    def __lt__(self, other):
        # Ta funkcja mówi tabeli, jak porównywać dwa wiersze
        try:
            # Pobieramy surowe liczby z EditRole
            val1 = float(self.data(Qt.ItemDataRole.EditRole))
            val2 = float(other.data(Qt.ItemDataRole.EditRole))
            return val1 < val2
        except:
            # Jeśli coś pójdzie nie tak, sortuj normalnie
            return super().__lt__(other)


# --- WORKER ---
class FirebaseSyncWorker(QThread):
    data_ready = pyqtSignal(dict, dict)
    error_occurred = pyqtSignal(str)
    status_update = pyqtSignal(str)

    def run(self):
        self.status_update.emit("⏳ Pobieranie...")
        try:
            r_addr = requests.get(f"{FIREBASE_URL}adresy.json", timeout=10)
            addr = r_addr.json() if r_addr.status_code == 200 and r_addr.json() else {}

            r_routes = requests.get(f"{FIREBASE_URL}trasy.json", timeout=10)
            routes = r_routes.json() if r_routes.status_code == 200 and r_routes.json() else {}

            self.data_ready.emit(addr, routes)
        except Exception as e:
            self.error_occurred.emit(str(e))


# --- GŁÓWNA APLIKACJA ---
class LogisticsAppFinalV16(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("WIP")
        self.setMinimumSize(1300, 850)

        # --- DANE ---
        self.raw_data = []
        self.stats = {}
        self.saved_addresses = {}
        self.routes_db = {}
        self.current_view = "all"
        self.current_columns = []

        self.product_mapping = {
            'W': 'ESU', 'U': 'ECX', 'P': 'WPX', 'H': 'ESI',
            'C': 'CMX', 'T': 'TDT', 'K': 'TDK', 'N': 'DOM',
            'Y': 'TDY', 'Q': 'WMX', "D" : "DOX"
        }

        self.init_ui()
        self.apply_styles()
        self.start_sync()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(15, 15, 15, 15)

        # 1. HEADER
        header = QFrame()
        header.setObjectName("Header")
        h_layout = QHBoxLayout(header)

        lbl_title = QLabel("LOGISTICS DASHBOARD")
        lbl_title.setObjectName("Title")

        self.btn_load = self.mk_btn("📂 Wczytaj Pliki", self.load_files)
        self.btn_sync = self.mk_btn("☁️ Odśwież Dane", self.start_sync)
        self.btn_export = self.mk_btn("💾 Eksportuj", self.export_reports)
        self.btn_export.setEnabled(False)

        self.lbl_status = QLabel("Gotowy")
        self.lbl_status.setObjectName("Status")

        h_layout.addWidget(lbl_title)
        h_layout.addStretch()
        h_layout.addWidget(self.btn_load)
        h_layout.addWidget(self.btn_sync)
        h_layout.addWidget(self.btn_export)
        h_layout.addStretch()
        h_layout.addWidget(self.lbl_status)
        layout.addWidget(header)

        # 2. FILTRY
        filters = QFrame()
        filters.setObjectName("Filters")
        f_layout = QHBoxLayout(filters)

        self.btn_all = self.mk_filter("📋 WSZYSTKO", "all")
        self.btn_multi = self.mk_filter("📦 WIELOPAKI", "multi")
        self.btn_pallet = self.mk_filter("🏗️ PALETY", "pallet")
        self.btn_twelve = self.mk_filter("⏰ DWUNASTKI", "twelve")

        f_layout.addWidget(self.btn_all)
        f_layout.addWidget(self.btn_multi)
        f_layout.addWidget(self.btn_pallet)
        f_layout.addWidget(self.btn_twelve)
        f_layout.addStretch()
        layout.addWidget(filters)

        # 3. TABELA
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.cellDoubleClicked.connect(self.on_double_click)

        layout.addWidget(self.table)

    def mk_btn(self, text, func):
        btn = QPushButton(text)
        btn.setFixedHeight(35)
        btn.clicked.connect(func)
        return btn

    def mk_filter(self, text, mode):
        btn = QPushButton(text)
        btn.setFixedHeight(35)
        btn.setCheckable(True)
        btn.setEnabled(False)
        btn.setProperty("filter", True)
        btn.clicked.connect(lambda: self.change_view(btn, mode))
        return btn

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #121212; }
            QWidget { color: #E0E0E0; font-family: 'Segoe UI'; font-size: 13px; }
            QFrame#Header, QFrame#Filters { background-color: #1E1E1E; border-radius: 8px; }
            QLabel#Title { color: #448AFF; font-size: 18px; font-weight: bold; }
            QPushButton { background-color: #263238; border: 1px solid #37474F; color: white; border-radius: 4px; padding: 0 15px; }
            QPushButton:hover { background-color: #37474F; }
            QPushButton[filter="true"]:checked { background-color: #FF6F00; border: none; }
            QTableWidget { background-color: #1E1E1E; border: none; }
            QHeaderView::section { background-color: #263238; color: white; padding: 8px; border: none; font-weight: bold; }
            QTableWidget::item { border-bottom: 1px solid #333; padding: 5px; }
            QTableWidget::item:selected { background-color: #2962FF; }
        """)

    # --- LOGIKA ---
    def load_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Pliki", "", "Excel/CSV (*.xlsx *.xls *.csv)")
        if not paths: return
        try:
            self.raw_data = []
            self.stats = {}
            for p in paths: self.process_single_file(p)
            for b in [self.btn_all, self.btn_multi, self.btn_pallet, self.btn_twelve, self.btn_export]: b.setEnabled(
                True)
            self.change_view(self.btn_all, "all")
            self.lbl_status.setText(f"Wierszy: {len(self.raw_data)}")
        except Exception as e:
            QMessageBox.critical(self, "Błąd", str(e))

    def get_headers(self, header_row):
        mapping = {}
        candidates = {
            'AWB': ['AWB', 'List przewozowy', 'HU Lvl 1', 'Master'],
            'Weight': ['Weight', 'Waga', 'Gross Weight', 'Masa'],
            'PieceID': ['Piece ID', 'PieceID', 'Index', 'Nr paczki'],
            'Product': ['Product', 'Produkt', 'Prod'],
            'ProdType': ['Prod Type', 'Typ produktu', 'Service'],
            'Receiver': ['Rcvr Name', 'Receiver', 'Odbiorca']
        }
        for idx, col_name in enumerate(header_row):
            if not col_name: continue
            col_name_clean = str(col_name).strip()
            for key, syn_list in candidates.items():
                for syn in syn_list:
                    if syn.lower() == col_name_clean.lower():
                        mapping[key] = idx
        return mapping

    def parse_float(self, val):
        if val is None: return 0.0
        s = str(val).replace(',', '.').strip()
        try:
            return float(s)
        except ValueError:
            return 0.0

    def process_single_file(self, path):
        rows = []
        if path.endswith('.csv'):
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                sample = f.read(1024)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;')
                except:
                    dialect = 'excel'
                reader = csv.reader(f, dialect)
                rows = list(reader)
        else:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True): rows.append(list(row))
            wb.close()

        if not rows: return
        current_map = self.get_headers(rows[0])
        if 'AWB' not in current_map: return

        for r in rows[1:]:
            def get_val(key):
                idx = current_map.get(key)
                return r[idx] if idx is not None and idx < len(r) else ""

            awb = str(get_val('AWB')).strip()
            if not awb: continue

            row_dict = {
                'AWB': awb,
                'Weight': self.parse_float(get_val('Weight')),
                'Product': str(get_val('Product')).strip().upper(),
                'ProdType': str(get_val('ProdType')).strip().upper(),
                'PieceID': str(get_val('PieceID')).strip(),
                'Receiver': str(get_val('Receiver')).strip()
            }
            self.raw_data.append(row_dict)

            if awb not in self.stats:
                self.stats[awb] = {'count': 0, 'total_weight': 0.0, 'product': row_dict['Product'],
                                   'receiver': row_dict['Receiver']}
            self.stats[awb]['count'] += 1
            self.stats[awb]['total_weight'] += row_dict['Weight']

    # --- WIDOKI I RENDEROWANIE ---
    def change_view(self, btn, mode):
        for b in [self.btn_all, self.btn_multi, self.btn_pallet, self.btn_twelve]: b.setChecked(False)
        btn.setChecked(True)
        self.show_view(mode)

    def show_view(self, mode):
        self.current_view = mode
        self.table.setSortingEnabled(False)
        self.table.clear()

        # DEFINICJA KOLUMN
        if mode == "all":
            cols = ["AWB", "Produkt", "Ilość", "Waga", "Odbiorca"]
        elif mode == "multi":
            cols = ["AWB", "Ilość", "Waga", "Produkt", "Odbiorca", "Adres", "Trasa"]
        elif mode == "pallet":
            cols = ["AWB", "Waga Paczki", "Odbiorca", "Adres", "Trasa"]
        elif mode == "twelve":
            cols = ["AWB", "PID", "Waga", "Produkt", "Odbiorca", "Adres"]

        self.current_columns = cols
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)

        # DANE
        data = []

        if mode == "all":
            for awb, stat in self.stats.items():
                data.append([
                    awb,
                    self.map_prod(stat['product']),
                    stat['count'],
                    stat['total_weight'],  # RAW FLOAT
                    stat['receiver']
                ])

        elif mode == "multi":
            grouped = {}
            for row in self.raw_data:
                if row['Weight'] > 30: continue
                norm = self.normalize_name(row['Receiver'])
                if not norm: continue
                if norm not in grouped:
                    grouped[norm] = {'recv': row['Receiver'], 'pcs': 0, 'w': 0.0, 'prod': set(), 'ref': row['AWB']}
                grouped[norm]['pcs'] += 1
                grouped[norm]['w'] += row['Weight']
                grouped[norm]['prod'].add(row['Product'])

            for d in grouped.values():
                if d['pcs'] >= 10 or d['w'] >= 100:
                    prod_s = self.map_prod(list(d['prod'])[0]) if len(d['prod']) == 1 else "MIX"
                    data.append([
                        d['ref'],
                        d['pcs'],
                        d['w'],  # RAW FLOAT
                        prod_s,
                        d['recv'],
                        self.get_addr(d['recv']),
                        self.routes_db.get(d['ref'], "-")
                    ])

        elif mode == "pallet":
            for row in self.raw_data:
                if row['Weight'] > 20:
                    data.append([
                        row['AWB'],
                        row['Weight'],  # RAW FLOAT
                        row['Receiver'],
                        self.get_addr(row['Receiver']),
                        self.routes_db.get(row['AWB'], "-")
                    ])

        elif mode == "twelve":
            for row in self.raw_data:
                if row['Product'] in ['C', 'Q'] or 'PRE' in row['ProdType']:
                    pid = ("J" + row['PieceID']) if row['PieceID'].startswith('JD') else ("00" + row['PieceID'])
                    data.append([
                        row['AWB'],
                        pid,
                        row['Weight'],  # RAW FLOAT
                        self.map_prod(row['Product']),
                        row['Receiver'],
                        self.get_addr(row['Receiver'])
                    ])

        # WSTAWIANIE DO TABELI
        self.table.setRowCount(len(data))

        for r_idx, row_val in enumerate(data):
            for c_idx, val in enumerate(row_val):

                # --- KLUCZOWE: UŻYCIE NumericItem DLA KOLUMN LICZBOWYCH ---
                header_name = cols[c_idx]
                is_numeric_col = header_name in ["Waga", "Waga Paczki", "Ilość"]

                if is_numeric_col and isinstance(val, (int, float)):
                    # UŻYWAMY SPECJALNEJ KLASY DO SORTOWANIA
                    item = NumericItem()
                    item.setData(Qt.ItemDataRole.EditRole, val)  # WARTOŚĆ DO SORTOWANIA

                    # FORMATOWANIE WYŚWIETLANIA
                    if isinstance(val, float):
                        item.setText(f"{val:.2f}")
                    else:
                        item.setText(str(val))

                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # WYŚRODKOWANIE

                else:
                    # STANDARDOWY ITEM DLA TEKSTU
                    item = QTableWidgetItem(str(val))

                    # Logika wyrównania dla tekstów
                    if header_name in ["Odbiorca", "Adres", "Receiver"]:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Środek dla AWB, Produkt, Trasa

                # Highlight dla dużych wielopaków
                if mode == "multi" and header_name == "Ilość" and isinstance(val, int) and val >= 50:
                    item.setBackground(QColor("#FF6F00"))
                    item.setForeground(QColor("white"))

                self.table.setItem(r_idx, c_idx, item)

        self.table.setSortingEnabled(True)

    # --- HELPERS ---
    def map_prod(self, c):
        return self.product_mapping.get(c, c)

    def normalize_name(self, n):
        return ''.join(e for e in n if e.isalnum()).upper() if n else ""

    def get_addr(self, recv):
        return self.saved_addresses.get(self.normalize_name(recv), "")

    # --- SYNCHRONIZACJA ---
    def start_sync(self):
        self.btn_sync.setEnabled(False)
        self.lbl_status.setText("Pobieranie...")
        self.worker = FirebaseSyncWorker()
        self.worker.data_ready.connect(self.on_sync_ok)
        self.worker.finished.connect(lambda: self.btn_sync.setEnabled(True))
        self.worker.start()

    def on_sync_ok(self, addr, routes):
        self.saved_addresses = addr
        self.routes_db = routes
        self.lbl_status.setText("Online")
        if self.raw_data: self.show_view(self.current_view)

    # --- INTERAKCJE ---
    def show_context_menu(self, pos):
        menu = QMenu()
        menu.addAction("📋 Kopiuj AWB", self.copy_awb)
        menu.addSeparator()
        menu.addAction("✏️ Edytuj Adres", self.edit_addr)
        menu.exec(self.table.mapToGlobal(pos))

    def copy_awb(self):
        rows = sorted(set(i.row() for i in self.table.selectedIndexes()))
        if rows and "AWB" in self.current_columns:
            txt = self.table.item(rows[0], self.current_columns.index("AWB")).text()
            QApplication.clipboard().setText(txt)

    def edit_addr(self):
        rows = sorted(set(i.row() for i in self.table.selectedIndexes()))
        if not rows: return

        col_name = "Odbiorca" if "Odbiorca" in self.current_columns else "Receiver"
        try:
            col = self.current_columns.index(col_name)
        except:
            return

        recv = self.table.item(rows[0], col).text()
        curr = self.get_addr(recv)
        new, ok = QInputDialog.getText(self, "Adres", f"Edytuj: {recv}", text=curr)
        if ok and new:
            norm = self.normalize_name(recv)
            self.saved_addresses[norm] = new
            from threading import Thread
            Thread(target=lambda: requests.patch(f"{FIREBASE_URL}adresy.json", json={norm: new}), daemon=True).start()
            self.show_view(self.current_view)

    def on_double_click(self, row, col):
        if self.current_view in ["multi", "pallet"] and self.current_columns[col] == "Trasa":
            awb = self.table.item(row, 0).text()
            item = self.table.item(row, col)
            nxt = "HEAVY 1" if item.text() in ["-", ""] else ("HEAVY 2" if item.text() == "HEAVY 1" else "-")

            self.routes_db[awb] = nxt
            item.setText(nxt)
            from threading import Thread
            Thread(target=lambda: requests.patch(f"{FIREBASE_URL}trasy.json", json={awb: nxt}), daemon=True).start()

    def export_reports(self):
        folder = QFileDialog.getExistingDirectory(self, "Folder")
        if not folder: return
        t_dir = os.path.join(folder, f"Raporty_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}")
        os.makedirs(t_dir, exist_ok=True)
        try:
            def save(n, h, d):
                wb = openpyxl.Workbook();
                ws = wb.active;
                ws.append(h)
                for c in ws[1]: c.font = Font(bold=True)
                for r in d: ws.append(r)
                wb.save(os.path.join(t_dir, n))

            multi, pallet, twelve = [], [], []
            gr = {}
            for r in self.raw_data:
                if r['Weight'] <= 30 and (norm := self.normalize_name(r['Receiver'])):
                    if norm not in gr: gr[norm] = {'p': 0, 'w': 0, 'pr': set(), 'ref': r['AWB'], 'n': r['Receiver']}
                    gr[norm]['p'] += 1;
                    gr[norm]['w'] += r['Weight'];
                    gr[norm]['pr'].add(r['Product'])
            for v in gr.values():
                if v['p'] >= 10 or v['w'] >= 100:
                    prod = self.map_prod(list(v['pr'])[0]) if len(v['pr']) == 1 else "MIX"
                    multi.append([v['ref'], v['p'], v['w'], prod, v['n'], self.get_addr(v['n']),
                                  self.routes_db.get(v['ref'], "-")])

            for r in self.raw_data:
                if r['Weight'] > 30:
                    pallet.append([r['AWB'], r['Weight'], r['Receiver'], self.get_addr(r['Receiver']),
                                   self.routes_db.get(r['AWB'], "-")])

            for r in self.raw_data:
                if r['Product'] in ['C', 'Q'] or 'PRE' in r['ProdType']:
                    pid = ("J" + r['PieceID']) if r['PieceID'].startswith('JD') else ("00" + r['PieceID'])
                    twelve.append([r['AWB'], pid, r['Weight'], self.map_prod(r['Product']), r['Receiver'],
                                   self.get_addr(r['Receiver'])])

            save("Wielopaki.xlsx", ["AWB", "Ilość", "Waga", "Produkt", "Odbiorca", "Adres", "Trasa"], multi)
            save("Palety.xlsx", ["AWB", "Waga Paczki", "Odbiorca", "Adres", "Trasa"], pallet)
            save("Dwunastki.xlsx", ["AWB", "PID", "Waga", "Produkt", "Odbiorca", "Adres"], twelve)
            QMessageBox.information(self, "Ok", f"Zapisano w: {t_dir}")
        except Exception as e:
            QMessageBox.critical(self, "Błąd", str(e))


if __name__ == "__main__":
    if os.name == 'nt':
        import ctypes

        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('log.master.16.0')
        except:
            pass
    app = QApplication(sys.argv)
    win = LogisticsAppFinalV16()
    win.show()
    sys.exit(app.exec())